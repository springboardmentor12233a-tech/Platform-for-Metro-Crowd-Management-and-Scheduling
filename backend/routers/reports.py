from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import StreamingResponse
from typing import Optional
from datetime import datetime, timedelta
import io
import csv
from bson import ObjectId
from backend.database import db_instance
from backend.auth import RoleChecker

router = APIRouter(prefix="/api/reports", tags=["Reports"])

analyst_only = RoleChecker(["Admin", "Analyst"])

@router.get("/generate")
async def generate_report(
    type: str = Query(..., pattern="^(passenger|station|occupancy|delay|peak_hour)$"),
    format: str = Query(..., pattern="^(pdf|xlsx|csv)$"),
    start_date: Optional[str] = Query(None),
    end_date: Optional[str] = Query(None),
    current_user: dict = Depends(analyst_only)
):
    db = db_instance.db
    if db is None:
        raise HTTPException(status_code=500, detail="Database offline")
        
    # Parse dates
    try:
        s_date = datetime.strptime(start_date, "%Y-%m-%d") if start_date else datetime.utcnow() - timedelta(days=7)
        e_date = datetime.strptime(end_date, "%Y-%m-%d") if end_date else datetime.utcnow()
    except ValueError:
        raise HTTPException(status_code=400, detail="Invalid date format. Use YYYY-MM-DD")
        
    # Fetch data based on type
    data = []
    headers = []
    filename = f"metroflow_{type}_report_{datetime.now().strftime('%Y%m%d')}"
    
    if type == "passenger":
        headers = ["Date/Time", "Station ID", "Station Name", "Passenger Count", "Inflow", "Outflow", "Waiting Passengers", "Crowd Level"]
        # Fetch crowd logs
        cursor = db.crowd_data.find({
            "timestamp": {"$gte": s_date, "$lte": e_date}
        }).sort("timestamp", -1).limit(100)
        
        async for c in cursor:
            # Get station name
            st_name = "Unknown Station"
            st = await db.stations.find_one({"_id": c["station_id"]})
            if st:
                st_name = st["name"]
            data.append([
                c["timestamp"].strftime("%Y-%m-%d %H:%M:%S"),
                str(c["station_id"]),
                st_name,
                c.get("passenger_count", 0),
                c.get("inflow", 0),
                c.get("outflow", 0),
                c.get("waiting_passengers", 0),
                c.get("crowd_level", "Green")
            ])
            
    elif type == "station":
        headers = ["Station Code", "Station Name", "Line", "Zone", "Platforms Count", "Coordinates", "Status"]
        cursor = db.stations.find({})
        async for s in cursor:
            data.append([
                s.get("station_code", ""),
                s.get("name", ""),
                s.get("line", ""),
                s.get("zone", ""),
                len(s.get("platforms", [])),
                f"{s.get('latitude', 0.0)}, {s.get('longitude', 0.0)}",
                s.get("status", "Active")
            ])
            
    elif type == "occupancy":
        headers = ["Train Number", "Train Name", "Line/Route", "Capacity", "Current Occupancy", "Occupancy Pct", "Status"]
        cursor = db.trains.find({})
        async for t in cursor:
            # Get route name
            r_name = "Unknown Route"
            route = await db.routes.find_one({"_id": t["route_id"]})
            if route:
                r_name = route["name"]
            occ_pct = round((t.get("current_occupancy", 0) / t.get("capacity", 1)) * 100, 1)
            data.append([
                t.get("train_number", ""),
                t.get("train_name", ""),
                r_name,
                t.get("capacity", 0),
                t.get("current_occupancy", 0),
                f"{occ_pct}%",
                t.get("status", "")
            ])
            
    elif type == "delay":
        headers = ["Date", "Train", "Station", "Platform", "Scheduled Time", "Actual Time", "Delay (Mins)", "Status"]
        cursor = db.schedules.find({
            "status": "Delayed"
        }).limit(100)
        async for s in cursor:
            t = await db.trains.find_one({"_id": s["train_id"]})
            st = await db.stations.find_one({"_id": s["station_id"]})
            data.append([
                datetime.now().strftime("%Y-%m-%d"), # Mock date
                t.get("train_number", "Unknown") if t else "Unknown",
                st.get("name", "Unknown") if st else "Unknown",
                s.get("platform", 1),
                s.get("scheduled_arrival", ""),
                s.get("actual_arrival", ""),
                s.get("delay_min", 0),
                s.get("status", "")
            ])
            
    elif type == "peak_hour":
        headers = ["Time Period", "Avg Inflow", "Avg Outflow", "Avg Passengers", "Congestion Indicator"]
        # Aggregated peak periods based on dummy historical insights
        data = [
            ["Early Morning (06:00-08:00)", 80, 75, 155, "Low"],
            ["Morning Peak (08:00-10:00)", 450, 420, 870, "Critical"],
            ["Mid Day (10:00-16:00)", 180, 195, 375, "Moderate"],
            ["Evening Peak (16:00-19:30)", 510, 480, 990, "Critical"],
            ["Late Night (19:30-23:00)", 95, 110, 205, "Low"]
        ]
        
    # ----------------------------------------------------
    # FORMAT EXPORTERS
    # ----------------------------------------------------
    if format == "csv":
        output = io.StringIO()
        writer = csv.writer(output)
        writer.writerow(headers)
        writer.writerows(data)
        output.seek(0)
        return StreamingResponse(
            io.BytesIO(output.getvalue().encode("utf-8")),
            media_type="text/csv",
            headers={"Content-Disposition": f"attachment; filename={filename}.csv"}
        )
        
    elif format == "xlsx":
        try:
            import openpyxl
            from openpyxl.styles import Font, Alignment, PatternFill
        except ImportError:
            raise HTTPException(status_code=500, detail="openpyxl is not installed")
            
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = f"{type.capitalize()} Report"
        
        # Style headers
        header_fill = PatternFill(start_color="1F2937", end_color="1F2937", fill_type="solid")
        header_font = Font(name="Arial", size=11, bold=True, color="FFFFFF")
        center_align = Alignment(horizontal="center", vertical="center")
        
        ws.append(headers)
        for col_num in range(1, len(headers) + 1):
            cell = ws.cell(row=1, column=col_num)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center_align
            
        # Add data
        for row in data:
            ws.append(row)
            
        # Auto-adjust column width
        for col in ws.columns:
            max_len = max(len(str(cell.value or '')) for cell in col)
            col_letter = openpyxl.utils.get_column_letter(col[0].column)
            ws.column_dimensions[col_letter].width = max(max_len + 3, 10)
            
        file_stream = io.BytesIO()
        wb.save(file_stream)
        file_stream.seek(0)
        return StreamingResponse(
            file_stream,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}.xlsx"}
        )
        
    elif format == "pdf":
        try:
            from reportlab.lib.pagesizes import letter
            from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
            from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
            from reportlab.lib import colors
        except ImportError:
            raise HTTPException(status_code=500, detail="reportlab is not installed")
            
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=letter, rightMargin=36, leftMargin=36, topMargin=36, bottomMargin=36)
        elements = []
        
        styles = getSampleStyleSheet()
        # Custom Title Style
        title_style = ParagraphStyle(
            name="TitleStyle",
            parent=styles["Heading1"],
            fontSize=18,
            leading=22,
            textColor=colors.HexColor("#1E3A8A"), # Navy Blue
            spaceAfter=15
        )
        
        meta_style = ParagraphStyle(
            name="MetaStyle",
            parent=styles["Normal"],
            fontSize=9,
            textColor=colors.HexColor("#4B5563"),
            spaceAfter=20
        )
        
        # Add Title
        elements.append(Paragraph(f"AI MetroFlow - {type.replace('_', ' ').title()} Report", title_style))
        elements.append(Paragraph(f"Generated on: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')} | Scope: {s_date.strftime('%Y-%m-%d')} to {e_date.strftime('%Y-%m-%d')}", meta_style))
        elements.append(Spacer(1, 10))
        
        # Table formatting
        table_data = [headers] + [[str(item) for item in row] for row in data]
        
        # We need to wrap cells in Paragraph to support auto-wrapping inside tables
        cell_style = ParagraphStyle(name="Cell", parent=styles["Normal"], fontSize=8, leading=10)
        table_paragraphs = []
        for r_idx, row in enumerate(table_data):
            row_paragraphs = []
            for col_idx, item in enumerate(row):
                if r_idx == 0:
                    row_paragraphs.append(Paragraph(f"<b>{item}</b>", ParagraphStyle(name="HCell", parent=cell_style, textColor=colors.white, fontSize=9)))
                else:
                    row_paragraphs.append(Paragraph(item, cell_style))
            table_paragraphs.append(row_paragraphs)
            
        # Table Styling
        pdf_table = Table(table_paragraphs, colWidths=[(letter[0] - 72) / len(headers)] * len(headers))
        pdf_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor("#1F2937")), # Dark gray header
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('VALIGN', (0, 0), (-1, -1), 'TOP'),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor("#E5E7EB")), # Light grid
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor("#F9FAFB")]), # Alternating rows
            ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
            ('TOPPADDING', (0, 0), (-1, -1), 6),
        ]))
        
        elements.append(pdf_table)
        doc.build(elements)
        buffer.seek(0)
        
        return StreamingResponse(
            buffer,
            media_type="application/pdf",
            headers={"Content-Disposition": f"attachment; filename={filename}.pdf"}
        )


@router.get("/traffic-report")
async def get_traffic_report(
    format: str = Query("csv", pattern="^(pdf|xlsx|csv)$"),
    start_date: Optional[str] = Query(None),
    end_date: Optional[str] = Query(None),
    current_user: dict = Depends(analyst_only)
):
    return await generate_report("passenger", format, start_date, end_date, current_user)


@router.get("/frequency-report")
async def get_frequency_report(
    format: str = Query("csv", pattern="^(pdf|xlsx|csv)$"),
    start_date: Optional[str] = Query(None),
    end_date: Optional[str] = Query(None),
    current_user: dict = Depends(analyst_only)
):
    return await generate_report("occupancy", format, start_date, end_date, current_user)


# ============================================================
# DATASET-POWERED ANALYTICS ENDPOINTS
# Pre-computed from delhi_metro_updated.csv, public_transport_delays.csv
# ============================================================
import os
import pandas as pd
import numpy as np
from functools import lru_cache

_ANALYTICS_CACHE = {}

def _load_analytics_cache():
    """Load and precompute analytics data from datasets on first call."""
    global _ANALYTICS_CACHE
    if _ANALYTICS_CACHE:
        return _ANALYTICS_CACHE

    metro_path = "datasets/delhi_metro_updated.csv"
    delay_path = "datasets/public_transport_delays.csv"
    network_path = "datasets/Delhi-Metro-Network.csv"

    result = {}

    # --- 1. Passenger Monthly Trends ---
    if os.path.exists(metro_path):
        try:
            df = pd.read_csv(metro_path, parse_dates=['Date'])
            df['From_Station'] = df['From_Station'].str.strip()
            df['To_Station'] = df['To_Station'].str.strip()
            df['YearMonth'] = df['Date'].dt.to_period('M').astype(str)
            monthly = df.groupby('YearMonth')['Passengers'].agg(['sum', 'mean', 'count']).reset_index()
            monthly.columns = ['month', 'total_passengers', 'avg_per_trip', 'total_trips']
            monthly = monthly.sort_values('month')
            result['passenger_trends'] = monthly.to_dict(orient='records')

            # --- 2. Top Routes ---
            df['route'] = df['From_Station'] + ' → ' + df['To_Station']
            top_routes = df.groupby('route').agg(
                total_passengers=('Passengers', 'sum'),
                total_trips=('TripID', 'count'),
                avg_fare=('Fare', 'mean'),
                avg_distance=('Distance_km', 'mean')
            ).reset_index()
            top_routes = top_routes.nlargest(10, 'total_passengers')
            result['top_routes'] = top_routes.round(2).to_dict(orient='records')

            # --- 3. Ticket type breakdown ---
            ticket_types = df.groupby('Ticket_Type')['Passengers'].sum().reset_index()
            ticket_types.columns = ['ticket_type', 'passengers']
            result['ticket_breakdown'] = ticket_types.to_dict(orient='records')

            # --- 4. Day of week patterns ---
            df['DayOfWeek'] = df['Date'].dt.day_name()
            day_order = ['Monday','Tuesday','Wednesday','Thursday','Friday','Saturday','Sunday']
            dow = df.groupby('DayOfWeek')['Passengers'].agg(['sum','mean']).reindex(day_order).reset_index()
            dow.columns = ['day', 'total_passengers', 'avg_passengers']
            result['day_patterns'] = dow.round(1).to_dict(orient='records')
        except Exception as e:
            print(f"Analytics cache: metro CSV error: {e}")

    # --- 5. Delay Factors ---
    if os.path.exists(delay_path):
        try:
            df2 = pd.read_csv(delay_path)
            metro_delays = df2[df2['transport_type'] == 'Metro'].copy()
            if len(metro_delays) < 50:
                metro_delays = df2.copy()

            # Weather impact
            weather_impact = metro_delays.groupby('weather_condition').agg(
                delayed_rate=('delayed', 'mean'),
                avg_delay_min=('actual_arrival_delay_min', 'mean'),
                count=('trip_id', 'count')
            ).reset_index()
            weather_impact.columns = ['weather', 'delayed_rate', 'avg_delay_min', 'count']
            result['weather_impact'] = weather_impact.round(2).to_dict(orient='records')

            # Event type impact
            df2_events = df2.copy()
            df2_events['event_type'] = df2_events['event_type'].fillna('No Event')
            event_impact = df2_events.groupby('event_type').agg(
                delayed_rate=('delayed', 'mean'),
                avg_delay_min=('actual_arrival_delay_min', 'mean')
            ).reset_index()
            event_impact.columns = ['event_type', 'delayed_rate', 'avg_delay_min']
            result['event_impact'] = event_impact.round(2).to_dict(orient='records')

            # Season pattern
            season_impact = df2.groupby('season').agg(
                delayed_rate=('delayed', 'mean'),
                avg_delay_min=('actual_arrival_delay_min', 'mean')
            ).reset_index()
            season_impact.columns = ['season', 'delayed_rate', 'avg_delay_min']
            result['season_impact'] = season_impact.round(2).to_dict(orient='records')
        except Exception as e:
            print(f"Analytics cache: delay CSV error: {e}")

    # --- 6. Network overview ---
    if os.path.exists(network_path):
        try:
            df3 = pd.read_csv(network_path)
            df3['Line'] = df3['Line'].str.strip()
            line_stats = df3.groupby('Line').agg(
                station_count=('Station Name', 'count'),
                total_km=('Distance from Start (km)', 'max')
            ).reset_index()
            line_stats.columns = ['line', 'station_count', 'total_km']
            result['network_overview'] = line_stats.sort_values('station_count', ascending=False).to_dict(orient='records')
        except Exception as e:
            print(f"Analytics cache: network CSV error: {e}")

    _ANALYTICS_CACHE = result
    return result


@router.get("/analytics/passenger-trends")
async def get_passenger_trends(current_user: dict = Depends(analyst_only)):
    """Monthly passenger trends from real delhi_metro_updated.csv dataset."""
    data = _load_analytics_cache()
    return {
        "trends": data.get('passenger_trends', []),
        "day_patterns": data.get('day_patterns', []),
        "ticket_breakdown": data.get('ticket_breakdown', []),
        "source": "delhi_metro_updated.csv (150,000 trip records)"
    }


@router.get("/analytics/top-routes")
async def get_top_routes(current_user: dict = Depends(analyst_only)):
    """Top 10 busiest origin-destination routes from real dataset."""
    data = _load_analytics_cache()
    return {
        "routes": data.get('top_routes', []),
        "source": "delhi_metro_updated.csv (150,000 trip records)"
    }


@router.get("/analytics/delay-factors")
async def get_delay_factors(current_user: dict = Depends(analyst_only)):
    """Weather, event, and seasonal delay analysis from public_transport_delays.csv."""
    data = _load_analytics_cache()
    return {
        "weather_impact": data.get('weather_impact', []),
        "event_impact": data.get('event_impact', []),
        "season_impact": data.get('season_impact', []),
        "source": "public_transport_delays.csv (2,000 records)"
    }


@router.get("/analytics/network-overview")
async def get_network_overview(current_user: dict = Depends(analyst_only)):
    """Delhi Metro network line statistics from Delhi-Metro-Network.csv."""
    data = _load_analytics_cache()
    return {
        "lines": data.get('network_overview', []),
        "source": "Delhi-Metro-Network.csv (285 stations)"
    }
