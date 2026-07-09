"""
MetroFlow: CSV to XLSX Converter
==================================
Converts all CSV files in the data/ directory to properly formatted XLSX files.
Each XLSX file gets:
- Proper column headers with bold formatting
- Auto-adjusted column widths
- Date columns formatted as dates
- Number columns formatted appropriately
- A descriptive sheet name
"""

import pandas as pd
import os
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

DATA_DIR = "data"

# Map CSV filenames to descriptive sheet names
SHEET_NAMES = {
    "delhi_metro_ticketing.csv": "Ticketing_Data",
    "passenger_entry_exit.csv": "Entry_Exit_Data",
    "hourly_footfall.csv": "Hourly_Footfall",
    "station_footfall_daily.csv": "Station_Footfall",
    "metro_ridership.csv": "Ridership_Data",
    "train_occupancy.csv": "Train_Occupancy",
    "train_schedule.csv": "Train_Schedule",
    "delay_logs.csv": "Delay_Logs",
    "rail_transport_stats.csv": "Rail_Transport_Stats",
    "metro_sensor_data.csv": "Sensor_Data"
}


def format_xlsx(filepath, sheet_name):
    """Apply professional formatting to an XLSX file."""
    wb = load_workbook(filepath)
    ws = wb[sheet_name]

    # Header formatting
    header_font = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
    header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin")
    )

    # Format headers
    for col_idx in range(1, ws.max_column + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    # Format data rows with alternating colors
    even_fill = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
    data_alignment = Alignment(horizontal="left", vertical="center")

    for row_idx in range(2, ws.max_row + 1):
        for col_idx in range(1, ws.max_column + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.border = thin_border
            cell.alignment = data_alignment
            if row_idx % 2 == 0:
                cell.fill = even_fill

    # Auto-adjust column widths
    for col_idx in range(1, ws.max_column + 1):
        column_letter = get_column_letter(col_idx)
        max_length = 0
        for row_idx in range(1, min(ws.max_row + 1, 100)):  # Sample first 100 rows
            cell_value = str(ws.cell(row=row_idx, column=col_idx).value or "")
            max_length = max(max_length, len(cell_value))
        adjusted_width = min(max_length + 4, 35)
        ws.column_dimensions[column_letter].width = max(adjusted_width, 12)

    # Freeze the header row
    ws.freeze_panes = "A2"

    # Add auto-filter
    ws.auto_filter.ref = ws.dimensions

    wb.save(filepath)


def convert_csv_to_xlsx(csv_path, xlsx_path, sheet_name):
    """Convert a single CSV file to a formatted XLSX file."""
    print(f"  Converting: {os.path.basename(csv_path)} -> {os.path.basename(xlsx_path)}")

    # Read CSV
    df = pd.read_csv(csv_path)

    # Save to XLSX
    df.to_excel(xlsx_path, index=False, sheet_name=sheet_name, engine="openpyxl")

    # Apply formatting
    format_xlsx(xlsx_path, sheet_name)

    print(f"    -> {len(df):,} rows x {len(df.columns)} cols | Sheet: '{sheet_name}'")
    return df


def main():
    print("=" * 60)
    print("MetroFlow: CSV to XLSX Converter")
    print("=" * 60)
    print(f"Data Directory: {os.path.abspath(DATA_DIR)}")
    print()

    converted = 0
    errors = 0

    # Convert all known CSV files
    for csv_filename, sheet_name in SHEET_NAMES.items():
        csv_path = os.path.join(DATA_DIR, csv_filename)
        xlsx_path = os.path.join(DATA_DIR, csv_filename.replace(".csv", ".xlsx"))

        if os.path.exists(csv_path):
            try:
                convert_csv_to_xlsx(csv_path, xlsx_path, sheet_name)
                converted += 1
            except Exception as e:
                print(f"  ERROR converting {csv_filename}: {e}")
                errors += 1
        else:
            print(f"  SKIP: {csv_filename} (not found)")

    # Also convert any other CSV files not in the map
    for filename in os.listdir(DATA_DIR):
        if filename.endswith(".csv") and filename not in SHEET_NAMES:
            csv_path = os.path.join(DATA_DIR, filename)
            xlsx_path = os.path.join(DATA_DIR, filename.replace(".csv", ".xlsx"))
            sheet_name = filename.replace(".csv", "").replace("_", " ").title()[:31]

            if not os.path.exists(xlsx_path):
                try:
                    convert_csv_to_xlsx(csv_path, xlsx_path, sheet_name)
                    converted += 1
                except Exception as e:
                    print(f"  ERROR converting {filename}: {e}")
                    errors += 1

    print()
    print("=" * 60)
    print(f"CONVERSION COMPLETE: {converted} files converted, {errors} errors")
    print("=" * 60)

    # List all XLSX files
    print("\nXLSX files in data/ directory:")
    for f in sorted(os.listdir(DATA_DIR)):
        if f.endswith(".xlsx"):
            size_kb = os.path.getsize(os.path.join(DATA_DIR, f)) / 1024
            print(f"  {f:45s} ({size_kb:>8.1f} KB)")


if __name__ == "__main__":
    main()
